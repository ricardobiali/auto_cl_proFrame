import json
from pathlib import Path
import sys
import os
import time
import openpyxl
from openpyxl.styles import numbers
import pandas as pd
import win32com.client

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
from backend.sap_manager.sap_connect import get_sap_free_session, start_sap_manager, start_connection

from backend.sap_manager.ysrelcont import executar_ysrelcont
from backend.sap_manager.ko03 import executar_ko03
from backend.sap_manager.ks13 import executar_ks13


# =========================================================
# ✅ ALTERAÇÃO MÍNIMA: requests.json agora vem do AppData
# =========================================================
APP_NAME = "AUTO_CL"

def _requests_path_appdata() -> Path:
    """
    requests.json persistente em AppData:
    %LOCALAPPDATA%\\AUTO_CL\\requests.json
    """
    base = os.environ.get("LOCALAPPDATA")
    if base:
        appdata_dir = Path(base) / APP_NAME
    else:
        appdata_dir = Path.home() / f".{APP_NAME.lower()}"
    appdata_dir.mkdir(parents=True, exist_ok=True)
    return appdata_dir / "requests.json"


# Caminho do requests.json (centralizado em AppData)
requests_path = _requests_path_appdata()

if not requests_path.exists():
    print(f"[ERRO] Arquivo requests.json não encontrado em: {requests_path}")
    sys.exit(1)

# Lê o arquivo JSON
try:
    with open(requests_path, "r", encoding="utf-8") as f:
        data = json.load(f)
except Exception as e:
    print(f"[ERRO] Falha ao ler requests.json ({requests_path}): {e}")
    sys.exit(1)


# --- NOVO BLOCO: monta lista de arquivos a processar ---
files_reduzida = []

# Caso o JSON siga o novo formato com bloco "destino"
if "destino" in data:
    destino_block = data.get("destino", [])
    if not isinstance(destino_block, list):
        destino_block = [destino_block]

    for destino_dict in destino_block:
        if not isinstance(destino_dict, dict):
            continue
        for key in sorted(
            destino_dict.keys(),
            key=lambda x: int(x.replace("file_completa", "")) if x.startswith("file_completa") and x != "file_completa" else 0
        ):
            path_file = destino_dict.get(key)
            if path_file and os.path.exists(path_file):
                files_reduzida.append(path_file)
else:
    # Caso antigo: usa apenas o file_reduzida
    if "file_reduzida" in data and os.path.exists(data["file_reduzida"]):
        files_reduzida.append(data["file_reduzida"])

# Extrai path3 do bloco "paths"
path3_value = ""
if "paths" in data and isinstance(data["paths"], list) and len(data["paths"]) > 0:
    path3_value = (data["paths"][0].get("path3", "") or "").strip()

if not path3_value:
    print("[ERRO] 'path3' vazio no requests.json (paths[0].path3).")
    sys.exit(1)

if not files_reduzida:
    print("[ERRO] Nenhum arquivo válido encontrado para processar (destino/file_reduzida).")
    print("status_error")
    sys.exit(1)

# --- Processa cada arquivo da lista em sequência ---
for path_origin in files_reduzida:

    # --- Caminhos ---
    arquivo_origem = Path(path_origin)
    pasta_destino = Path(path3_value)
    os.makedirs(pasta_destino, exist_ok=True)

    # --- Lê o arquivo fonte ---
    try:
        df = pd.read_csv(arquivo_origem, sep=';', encoding='utf-8', low_memory=False)
    except UnicodeDecodeError:
        df = pd.read_csv(arquivo_origem, sep=';', encoding='latin1', low_memory=False)

    # --- Lista de colunas desejadas ---
    colunas_desejadas = [
        "Identificação DrillD", "Nº documento", "Linha lçto.", "Empresa", "Exercício", "Período",
        "Trimestre/Ano", "Data lçto.", "Data documento", "Nº doc.referên.", "Denominação", "Txt.cab.doc.",
        "Def.projeto", "Den. do projeto", "Elemento PEP", "Denominação do PEP", "Objeto", "Atividade Petrobras",
        "Descrição Ativ", "Cta.contrapart.", "Denom.conta contrap.", "Centro", "Cen.cst.solic.", "Centro de lucro",
        "Classe de custo", "Tp.doc.", "Desc Classe de Custo", "Valor/Moeda obj", "Moeda do objeto",
        "Valor total em reais", "Val suj cont loc R$", "Valor cont local R$", "Valor/moeda ACC", "Moeda da ACC",
        "Moeda transação", "Objeto parceiro", "Denom.obj.parc.", "Material", "Denominação", "Doc.compras",
        "Trat. Cont. Local", "MIGO", "MIRO", "Perc Cont Local Calc", "Certificado C.L.", "Perc Cont Local Info",
        "Justificativa %", "Taxa câmbio", "Grp.class.custo", "Doc.de estorno", "Doc.estornado", "Descrição da linha",
        "Código Regra", "Nat. G. Cal", "Descrição calculada", "Reclassificação", "Fase Consolidada", "Nat. Gast. Cons",
        "Perc Cont Local Con.", "Descrição con.", "Protocolo", "CNPJ do fornecedor", "Data Doc. Fiscal", "Referência",
        "Valor Total NF Reais", "Nº NF", "Nº da NF-e", "Doc.material", "It.  Material", "Tipo avaliação",
        "Código campo/bloco", "Sigla campo/bloco", "Contrato", "Forn. pedido", "Tipo movimento", "Desc. forn. pedido",
        "Doc custo Expurgado", "Fator Apr.CCs Consol", "Código da unidade", "Tipo de Operação",
        "Denom.Tp.Operação", "Texto", "Sigla da Gerência", "Doc.faturamento", "Doc.ref.", "Prog Expl Obrig/Mín",
        "Denominação Obj.", "Status Item/pedido", "EAP Unica", "Ref.estorno", "Visão EAP ÚNICA",
        "Percent_Rateio_Jaz", "Vl Nacional Atual", "Nome  do Índice", "Mês/ano ref.", "Ft. correção"
    ]

    # --- Verifica colunas existentes ---
    colunas_existentes = [c for c in colunas_desejadas if c in df.columns]
    colunas_faltando = [c for c in colunas_desejadas if c not in df.columns]

    if colunas_faltando:
        print("As seguintes colunas não foram encontradas no arquivo:")
        for c in colunas_faltando:
            print("  -", c)

    # --- Cria DataFrame reduzido ---
    df_reduzido = df[colunas_existentes].copy()

    # --- Remove linhas com 'X' em 'Doc custo Expurgado' ---
    if "Doc custo Expurgado" in df_reduzido.columns:
        linhas_antes = len(df_reduzido)
        df_reduzido = df_reduzido[
            df_reduzido["Doc custo Expurgado"].astype(str).str.strip().str.upper() != "X"
        ]
        print(f"{linhas_antes - len(df_reduzido)} linhas removidas (Doc custo Expurgado = 'X').")

    def sap_str_para_float(valor):
        if pd.isna(valor):
            return 0.0

        s = str(valor).strip()
        if not s:
            return 0.0

        # SAP: negativo vem no final (ex: 1,000.00-)
        negativo = s.endswith("-")
        s = s.replace("-", "")

        # Remove separador de milhar
        s = s.replace(",", "")

        try:
            num = float(s)
            return -num if negativo else num
        except ValueError:
            return 0.0

    # --- Converter e formatar colunas numéricas no padrão brasileiro ---
    colunas_numericas = [
        "Valor/Moeda obj",
        "Valor total em reais",
        "Val suj cont loc R$",
        "Valor cont local R$",
        "Valor/moeda ACC",
    ]

    for col in colunas_numericas:
        if col in df_reduzido.columns:
            df_reduzido[col] = df_reduzido[col].apply(sap_str_para_float)

    # --- Criar coluna "Estrangeiro $" logo após "Valor/moeda ACC" ---
    if "Val suj cont loc R$" in df_reduzido.columns and "Valor cont local R$" in df_reduzido.columns:
        idx_acc = df_reduzido.columns.get_loc("Valor cont local R$") + 1
        df_reduzido.insert(idx_acc, "Estrangeiro $", df_reduzido["Val suj cont loc R$"] - df_reduzido["Valor cont local R$"])

    # --- Função de formatação brasileira ---
    def formata_brasileiro(x):
        if pd.notnull(x):
            return f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return x

    # --- Aplica formatação ---
    for col in colunas_numericas + ["Estrangeiro $"]:
        if col in df_reduzido.columns:
            df_reduzido[col] = df_reduzido[col].apply(formata_brasileiro)

    # --- Adicionar colunas vazias ---
    novas_colunas = ["Tipo de Gasto", "Bem/Serviço", "Gestor do Contrato", "Gerência responsável pelo objeto parceiro", "Disciplina"]
    for col in novas_colunas:
        df_reduzido[col] = ""

    # --- Preencher Tipo de Gasto ---
    if "Tipo de Gasto" in df_reduzido.columns:
        # Garante que as colunas necessárias existam
        for col in ["Protocolo", "Objeto parceiro", "Doc.material"]:
            if col not in df_reduzido.columns:
                df_reduzido[col] = None

        # Converte tipos
        protocolo_num = pd.to_numeric(df_reduzido["Protocolo"], errors='coerce').fillna(0)
        doc_material_num = pd.to_numeric(df_reduzido["Doc.material"], errors='coerce').fillna(0)

        # Limpa e normaliza Objeto parceiro
        objeto_parceiro = (
            df_reduzido["Objeto parceiro"]
            .astype(str)
            .str.strip()
            .fillna("")  # substitui NaN por string vazia
        )
        # remove casos em que virou "nan" como texto
        objeto_parceiro = objeto_parceiro.replace("nan", "")

        # Começa tudo como 'Outros'
        df_reduzido["Tipo de Gasto"] = "Outros"

        mask_direto = protocolo_num > 0
        df_reduzido.loc[mask_direto, "Tipo de Gasto"] = "Direto"

        mask_indireto = (df_reduzido["Tipo de Gasto"] == "Outros") & (objeto_parceiro != "")
        df_reduzido.loc[mask_indireto, "Tipo de Gasto"] = "Indireto"

        mask_estoque = (
            (df_reduzido["Tipo de Gasto"] == "Outros")
            & (doc_material_num > 4899999999)
            & (doc_material_num < 5000000000)
        )
        df_reduzido.loc[mask_estoque, "Tipo de Gasto"] = "Estoque"

        print("Coluna 'Tipo de Gasto' preenchida conforme regras de prioridade (Direto Indireto Estoque Outros).")

    # --- Preencher Bem/Serviço ---
    if "Material" in df_reduzido.columns and "Bem/Serviço" in df_reduzido.columns:
        material_str = df_reduzido["Material"].astype(str).str.strip()

        df_reduzido.loc[
            material_str.str.match(r"^(50|70|80)"), "Bem/Serviço"
        ] = "Serviço"

        df_reduzido.loc[
            material_str.str.match(r"^(10|11|12)"), "Bem/Serviço"
        ] = "Material"

        print("Coluna 'Bem/Serviço' preenchida conforme prefixos de 'Material'.")
    else:
        print("Coluna 'Material' ou 'Bem/Serviço' não encontrada — nenhuma regra aplicada.")

    # Gerar lista única de contratos válidos (sem alterar o DataFrame)
    contratos_unicos = (
        df["Contrato"]
        .dropna()  # remove células vazias
        .astype(str)  # garante que tudo é string
        .str.strip()  # remove espaços em branco
        .str.replace(r"\.0$", "", regex=True)  # remove .0 no final
    )
    contratos_unicos = [
        c for c in contratos_unicos.unique() if c and c != "*"  # remove "*" e strings vazias
    ]

    # --- Inicialização SAP ---
    print("Iniciando SAP GUI...")
    start_sap_manager()
    start_connection()
    session = get_sap_free_session()
    time.sleep(2)

    # --- Executa transação SAP - Contratos/Gerentes ---
    print("Executando consulta YSRELCONT...")
    gerentes_por_contrato = executar_ysrelcont(session, contratos_unicos)
    if not isinstance(gerentes_por_contrato, dict):
        gerentes_por_contrato = {}
    print(f"Consulta SAP concluída. {len(gerentes_por_contrato)} contratos encontrados.")
    if not gerentes_por_contrato:
        print("YSRELCONT não retornou dados — Gestor do Contrato ficará em branco.")

    # --- Preenche coluna Contrato ---
    df_reduzido['Gestor do Contrato'] = (
        df_reduzido['Contrato']
        .astype(str)
        .str.strip()
        .str.replace(r"\.0$", "", regex=True)  # remove .0 caso exista
        .map(gerentes_por_contrato)
        .fillna('')
    )

    # Gerar lista única de objetos válidos (sem alterar o DataFrame)
    objetos_unicos = (
        df["Objeto parceiro"]
        .dropna()  # remove células vazias
        .astype(str)  # garante que tudo é string
        .str.strip()  # remove espaços em branco
    )
    objetos_unicos = [
        c for c in objetos_unicos.unique() if c and c != "*"  # remove "*" e strings vazias
    ]

    # --- Separa por tipo ---
    objetos_e = [o for o in objetos_unicos if o.startswith("E")]
    objetos_or = [o for o in objetos_unicos if o.startswith("OR")]

    # --- Execução KO03 + KS13 ---
    print("Executando KO03 (ordens OR - centros E)...")
    or_para_e = executar_ko03(session, objetos_or)
    print(f"{len(or_para_e)} ordens convertidas para centros de custo.")

    # Monta lista definitiva de objetos E
    objetos_definitivos = objetos_e + list(or_para_e.values())
    objetos_definitivos = list(set(objetos_definitivos))  # remove duplicatas

    print("Executando KS13 (centros E - gerências responsáveis)...")
    gerencias_por_objeto = executar_ks13(session, objetos_definitivos)
    print(f"{len(gerencias_por_objeto)} gerências encontradas.")

    # --- Preenche coluna no DataFrame ---
    def mapear_gerencia(obj):
        """Mapeia OR via seu E correspondente ou direto"""
        if pd.isna(obj):
            return ""
        obj = str(obj).strip()
        if obj.startswith("E"):
            return gerencias_por_objeto.get(obj, "")
        if obj.startswith("OR"):
            centro = or_para_e.get(obj, "")
            return gerencias_por_objeto.get(centro, "") if centro else ""
        return ""

    df_reduzido["Gerência responsável pelo objeto parceiro"] = df_reduzido["Objeto parceiro"].apply(mapear_gerencia)
    print("Coluna 'Gerência responsável pelo objeto parceiro' preenchida com sucesso.")
    # --- Fim do uso do SAP ---

    # --- Preencher coluna 'Disciplina' ---
    def aplica_regras_disciplina(df):
        def preenche_mascara(mascara, valor):
            df.loc[mascara & (df["Disciplina"] == ""), "Disciplina"] = valor

        # --- Função auxiliar para criar regex ---
        def prefixos_regex(lista):
            return "^(" + "|".join([p.replace("*", ".*") for p in lista]) + ")"

        # Dicionários de padrões
        diret_dict = {
            "LOEP": ([
                "LMS*", "US-LOG*", "US-SOEP*", "US-AP*", "LOEP*", "5331541*", "53337*", "53483*",
                "53531*", "53670*", "53671*", "536769*", "5367700*", "536771*", "536841*", "53684762",
                "53684763", "53684764", "53684765", "53684766", "53687*"
                ]),
            "POÇOS": ([
                "POCOS*", "CPM*", "EP-CPM*", "E&P-CPM*", "EPCPM*", "522*", "5237*", "52380*",
                "5239*", "529*", "529008*", "5294*", "5298*", "5309037*", "5309038*", "53090409",
                "5309041*", "5309042*", "53090430", "53090431", "53090432", "53090433", "5309045*",
                "5309046*", "53176439", "5317644*", "53315420", "53335*", "53485*", "534875*",
                "53561*", "53598*", "53626885", "53660*", "536695*", "536744*", "536755*", "536756*",
                "53676457", "53676458", "53684768", "53684769", "5370663*", "5370664*", "537585*",
                "537586*", "537589*"
                ]),
            "SUB": ([
                "E&P-SERV*", "SUB*", "IPSUB*", "500*", "52382*", "52388*", "5308*", "530902*", "5309047*",
                "5309048*", "530905*", "532*", "53315421", "53336*", "534879*", "5355*", "53564*", "53567*",
                "53592*", "53626886", "536694*", "536745*", "53679*", "53681*", "5370666*", "5370667*",
                "53739*"
                ]),
            "SRGE": ([
                "SH*", "SRGE*", "53535*", "53560*", "5357*", "5361*", "53625*", "536261*", "536262*",
                "536757*", "536758*", "5367644*", "53676450", "5367702*", "53682*", "536848*", "53685*",
                "53686*", "5369*", "53709*", "5372*", "53734*", "53737*", "53751*", "53756*", "537583*",
                "53759*"
                ]),
            "EXP": ([
                "EXP*", "AEXP*", "OEXP*", "508*", "510*", "512*", "52384*", "5309039*", "53090400",
                "53090401", "53090402", "53090403", "53090404", "53090405", "53090406", "53090407",
                "53090408", "53090434", "53090435", "53090436", "53090437", "53090438", "53090439",
                "53176435", "53176436", "53176437", "53176438"
                ]),
        }

        # Direto / Outros → coluna Sigla da Gerência
        for tipo in ["Direto", "Outros"]:
            mask_tipo = df["Tipo de Gasto"] == tipo
            for disc, prefixos in diret_dict.items():
                regex = prefixos_regex(prefixos)
                preenche_mascara(mask_tipo & df["Sigla da Gerência"].astype(str).str.match(regex, na=False), disc)

        # Indireto
        indireto = df["Tipo de Gasto"] == "Indireto"
        indireto_dict = {
            "LOEP": ["LMS*", "US-LOG*", "US-SOEP*", "US-AP*", "LOEP*"],
            "POÇOS": ["POCOS*", "CPM*", "EP-CPM*", "E&P-CPM*", "EPCPM*"],
            "SUB": ["E&P-SERV*", "SUB*", "IPSUB*"],
            "SRGE": ["SH*", "SRGE*"],
            "EXP": ["EXP*", "AEXP*", "OEXP*"],
        }
        for disc, prefixos in indireto_dict.items():
            regex = prefixos_regex(prefixos)
            preenche_mascara(indireto & df["Gerência responsável pelo objeto parceiro"].astype(str).str.match(regex, na=False), disc)

        # Indireto — Objeto parceiro
        op_dict = {
            "LOEP": ["E8*", "E9*"],
            "POÇOS": ["E5*", "E7*", "EI*", "EJ*", "EK*", "E000F41*", "E000F4Y*"],
            "SUB": ["E4*", "EY*", "EZ*", "E000GMN*"],
            "SRGE": ["SH*"],
        }
        for disc, prefixos in op_dict.items():
            regex = prefixos_regex(prefixos)
            preenche_mascara(indireto & df["Objeto parceiro"].astype(str).str.match(regex, na=False), disc)

        # Estoque — Código da unidade
        estoque = df["Tipo de Gasto"] == "Estoque"
        estoque_dict = {
            "POÇOS": ["PP00*", "PP01*", "PP03*", "PP04*", "PP05*", "PP07*", "PP08*", "PP09*"],
            "SRGE": ["PU01*", "PS01*"],
            "SUB": ["N100*", "PC01*", "PD00*", "PD03*", "PD04*", "PD05*", "PD08*", "PM04*", "PU03*", "PU43*"],
        }
        for disc, prefixos in estoque_dict.items():
            regex = prefixos_regex(prefixos)
            preenche_mascara(estoque & df["Código da unidade"].astype(str).str.match(regex, na=False), disc)

        # Preenche demais não contemplados
        df.loc[df["Disciplina"] == "", "Disciplina"] = "Demais"
        return df

    df_reduzido = aplica_regras_disciplina(df_reduzido)

    # --- Salvar arquivo final ---
    nome_base = os.path.basename(arquivo_origem)
    nome_reduzido = nome_base.replace(".txt", "_Reduzida.txt")
    caminho_saida = os.path.join(pasta_destino, nome_reduzido)
    nome_excel = Path(nome_reduzido).stem + ".xlsx"
    arquivo_excel = os.path.join(pasta_destino, nome_excel)

    # ✅ Salva o DataFrame reduzido antes de tentar reabrir
    df_reduzido.to_csv(caminho_saida, sep=";", index=False, encoding="utf-8")

# =========================================================
# BLOCO OPCIONAL – GERAÇÃO DE EXCEL
# Atualmente DESATIVADO por decisão de negócio
# Para reativar, basta remover os comentários abaixo
# =========================================================

    try:
        # # Lê o CSV e salva em Excel
        # df = pd.read_csv(caminho_saida, sep=";", encoding="utf-8", low_memory=False, dtype=str)

        # # Colunas que queremos converter para numérico (mesmo nome que você já usou)
        # colunas_formatar = [
        #     "Valor/Moeda obj",
        #     "Valor total em reais",
        #     "Val suj cont loc R$",
        #     "Valor cont local R$",
        #     "Valor/moeda ACC",
        #     "Estrangeiro $"
        # ]

        # # Converte as colunas formatadas em strings no padrão "1.234,56" para float 1234.56
        # for col in colunas_formatar:
        #     if col in df.columns:
        #         cleaned = (
        #             df[col]
        #             .astype(str)
        #             .str.replace(".", "", regex=False)
        #             .str.replace(",", ".", regex=False)
        #             .str.strip()
        #         )
        #         df[col] = pd.to_numeric(cleaned, errors="coerce")

        # # Salva o DataFrame (com colunas numéricas) direto em Excel
        # df.to_excel(arquivo_excel, index=False)

        # # Abre workbook e aplica formato numérico às colunas desejadas (se existirem)
        # wb = openpyxl.load_workbook(arquivo_excel)
        # ws = wb.active

        # # Mapeia cabeçalho -> índice (1-based)
        # header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

        # # Aplica formatação numérica (duas casas decimais)
        # for col in colunas_formatar:
        #     if col in header:
        #         col_idx = header[col]
        #         for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        #             for cell in row:
        #                 if isinstance(cell.value, (int, float)):
        #                     cell.number_format = '#,##0.00'

        # wb.save(arquivo_excel)

        status_done = "status_success"
        print(status_done)

    except Exception as e:
        status_done = "status_error"
        print(f"Ocorreu um erro: {e}")
        print(status_done)

print("\n Processamento reduzido concluído para todos os arquivos.")
